import pandas as pd
import re
import numpy as np
from pprint import pprint
import warnings
import openpyxl
import os
import variableUtils
import json
from sklearn.ensemble import RandomForestRegressor
from sklearn.impute import SimpleImputer
from sklearn.model_selection import train_test_split
from sklearn.metrics import mean_squared_error
from matplotlib import pyplot as plt
from IPython.display import display
from reportlab.lib.pagesizes import letter, landscape, A4, A3
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, PageBreak, Paragraph, Image, Spacer
from io import BytesIO
from reportlab.lib import colors
from matplotlib.backends.backend_pdf import PdfPages
from reportlab.platypus import Table as RLTable, TableStyle
from reportlab.platypus import Paragraph, Spacer, KeepTogether
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from PyPDF2 import PdfReader, PdfWriter
import datetime
from dateutil import parser
from itertools import combinations
# For data cleaning and preprocessing

def getFolderandFileName(filePath: str):
    """
    Gets the folder path and file name from a file path.

    Args:
        filePath (str): The path to the file.

    Returns:
        folderPath (str): The folder path containing the file.
        fileName (str): The name of the file.
    """
    folderPath, fileName = os.path.split(filePath)
    name, ext = os.path.splitext(fileName)
    return folderPath, name, ext

def convertDate(date_str):
    if isinstance(date_str, pd.Timestamp):
        return date_str.strftime('%d/%m/%Y')
    try:
        date_obj = parser.parse(date_str)
    except ValueError:
        print(f"Error parsing date: {date_str}, converting to datetime object")
        date_str = date_str.replace(' ', '')
        date_str = date_str.replace('th', '')
        date_str = date_str.replace('st', '')
        date_str = date_str.replace('nd', '')
        date_str = date_str.replace('rd', '')
        date_obj = parser.parse(date_str)
        print(f"Converted date: {date_obj}")
    except TypeError as e:
        print(f"TypeError: {e}")    
        date_obj = date_str
    print(f"Date: {date_obj}")
    return date_obj.strftime('%d/%m/%Y')

def loadWorkbook(filePath: str):
    """
    Loads an Excel workbook from the specified file path.
    
    Args:
        filePath (str): The path to the Excel file.
        
    Returns:
        workbook (openpyxl.Workbook): The loaded workbook object.
    """
    workbook = openpyxl.load_workbook(filePath, data_only=True)

    # Get folder path and file name
    folderPath, name, ext = getFolderandFileName(filePath)
    pprint(f"Loaded workbook: {folderPath} | {name} | {ext}")
    pprint(f"Workbook sheets: {workbook.sheetnames}")
    return workbook, folderPath, name

def getStudentList(listFile1 = 'data/Student ID for Kunal.xlsx', listFile2 = '2024/data/2024 MDS Student List_v10.xlsx', **kwargs):
    # Load the Excel file containing the student IDs
    # listDf1 = pd.read_excel(listFile1)
    # get cohort from kwargs
    cohort = kwargs.get('cohort', None)
    # if cohort is not None:
        # listDf1 = listDf1[listDf1['Cohort'] == cohort]
    # Get the student IDs as a list
    # studentList = list(listDf1[variableUtils.colId])

    # Load the Excel file containing the student IDs
    listDf2 = pd.read_excel(listFile2)
    if cohort is not None:
        listDf2 = listDf2[listDf2['Cohort'] == cohort]
    studentList = []
    # Get the student IDs as a list
    studentList += list(listDf2[variableUtils.colId])
    studentList = list(set(studentList))
    return studentList

def loadDfFromSheet(workbook, sheetName: str):
    """
    Load data from a specific sheet in a workbook into a pandas DataFrame.

    Parameters:
    - workbook: The workbook object containing the sheet.
    - sheetName: The name of the sheet to load data from.

    Returns:
    - df: A pandas DataFrame containing the data from the specified sheet.
    """

    sheet = workbook[sheetName]
    data = sheet.values  # returns a generator object
    columns = next(data)  # get the first row for the header
    df = pd.DataFrame(data, columns=columns)
    return df

def removeFakeNames(df: pd.DataFrame, colNameG: str, colNameF: str, studentList: list = None):
    """
    Removes fake names from the given columns in a DataFrame.

    Args:
        df (pandas.DataFrame): The DataFrame containing the columns.
        colNameG (str): The name of the column containing given names.
        colNameF (str): The name of the column containing family names.

    Returns:
        pandas.DataFrame: The DataFrame with fake names removed.
    """
    
    # Define a list of fake names to remove
    # fakeNames = ['Fake', 'Name', 'Test', 'Person', 'asdf', 'McTest']
    # # Filter out rows where both the given name and family name are fake
    # df = df[~df[colNameG].str.contains('|'.join(fakeNames), case=False, regex=False) & 
    #         ~df[colNameF].str.contains('|'.join(fakeNames), case=False, regex=False)]
    # Filter out the rows where the student ID is not in the student list
    if studentList is None:
        studentList = getStudentList()
    notInList = df[~df[variableUtils.colId].isin(studentList)]
    if len(notInList) > 0:
        for idx, row in notInList.iterrows():
            print(f"Student ID not in list: {row[variableUtils.colId]}, {row[colNameG]} {row[colNameF]}")
    df = df[df[variableUtils.colId].isin(studentList)]
    
    return df


def removeFirstRow(df: pd.DataFrame, keepHeader: bool = False):
    """
    Removes the first row from a DataFrame.

    Args:
        df (pandas.DataFrame): The DataFrame to remove the first row from.

    Returns:
        pandas.DataFrame: The DataFrame with the first row removed.
    """
    if keepHeader:
        newHeader = df.columns
    else:
        newHeader = df.iloc[0]

    df = df.iloc[1:]
    df.columns = newHeader
    df.reset_index(drop=True, inplace=True)
    return df

def getDfbyColumnValue(df: pd.DataFrame, colName: str, value: str):
    """
    Gets a subset of a DataFrame based on a specific column value.

    Args:
        df (pandas.DataFrame): The DataFrame to filter.
        colName (str): The name of the column to filter by.
        value (str): The value to filter by.

    Returns:
        pandas.DataFrame: The subset of the DataFrame where the column value matches the specified value.
    """
    dfTemp =  df[df[colName] == value]
    print(f"Found {len(dfTemp)} rows with {colName} = {value}")
    uniqueValues = dfTemp[colName].unique()
    counts = dfTemp[colName].value_counts()
    pprint(f"Unique values: {uniqueValues}")
    pprint(f"Value counts: {counts}")
    return dfTemp

def renameColumnsCode(df, renameDict={}):
    """
    Renames columns in a DataFrame based on specific conditions.

    Args:
        df (pandas.DataFrame): The DataFrame containing the columns to be renamed.

    Returns:
        pandas.DataFrame: The DataFrame with renamed columns.

    """
    counter_dict = {}  # Counter for each subject code and role MC column

    for col in df.columns:
        parts = re.split(r'\s+', col)
        code = parts[0]
        if len(parts) < 2:
            continue

        if code.isdigit():  # Check if the first part is a subject code and not Anesthesia or other columns
            # Identify specific tag (SIM, CLINIC)
            tag = f'({parts[1]})' if parts[1] in ['SIM', 'CLINIC'] else None
            role = 'student' if 'Student' in col else 'supervisor'

            # Formulate a unique key for counting based on subject code, tag, and role
            if tag:
                role_key = f"{code}_{role} {tag}"
            else:
                role_key = f"{code}_{role}"

            if role_key not in counter_dict:  # Initialize counter
                counter_dict[role_key] = 1
            else:  # Increment counter
                counter_dict[role_key] += 1

            # New name construction with tag
            new_name = f"{code}_MC{counter_dict[role_key]}_{role} {tag}" if tag else f"{code}_MC{counter_dict[role_key]}_{role}"
            renameDict[col] = new_name
        # change for anesthesia checklist columns
        elif 'Anesthesia' in col:
            # Use regex to split on any whitespace character and keep parts inside parentheses
            parts = re.split(r'\s-\s', col)
            checklist_part = parts[0]

            # Extract type from the checklist part e.g., "Infiltration", "Block"
            match = re.search(r'\((.*?)\)', checklist_part)
            anesthesia_type = match.group(1) if match else 'Unknown'

            # Identify role based on 'Student' or 'Supervisor'
            role = 'student' if 'Student' in col else 'supervisor'

            # Create a base for the new column name
            base_name = 'LA_'

            # Formulate a unique key for counting
            role_key = f"{base_name}_{anesthesia_type}_{role}"

            if role_key not in counter_dict:
                counter_dict[role_key] = 1
            else:
                counter_dict[role_key] += 1

            # New name construction
            new_name = f"{base_name}_MC{counter_dict[role_key]}_{role} ({anesthesia_type})"
            renameDict[col] = new_name

        elif 'Sharpening' in col:

            tag = f'({parts[1]})' if parts[1] in ['SIM', 'CLINIC'] else None
            role = 'student' if 'Student' in col else 'supervisor'

            # Create a base for the new column name
            base_name = 'Sharpening_'

            # Formulate a unique key for counting
            role_key = f"{base_name}_{role}"

            if role_key not in counter_dict:
                counter_dict[role_key] = 1
            else:
                counter_dict[role_key] += 1

            # New name construction
            new_name = f"{base_name}_MC{counter_dict[role_key]}_{role} ({tag})"
            renameDict[col] = new_name

        # else:
        #     print(f'Code: {code}, Not changing column name: {col}')

    # Use the rename dictionary to update column names
    df = df.rename(columns=renameDict)
    return df, renameDict

def renameColumnsOther(df, renameDict={}):
    print('Renaming Other columns')
    counter_dict = {}
    # Map for checklist categories to abbreviations
    checklist_map = {
        'Positioning and Ergonomics Checklist': 'PEC',
        'Infection control Checklist': 'ICC',
        'Record keeping Checklist': 'RKC',
        'Consent Checklist': 'CC'
    }
    for col in df.columns:
        # Identify the checklist category
        for checklist, abbreviation in checklist_map.items():
            if checklist in col:
                checklist_found = checklist
                checklist_abbr = abbreviation
                # print(f'Found checklist: {checklist_found}')
                break
        else:
            checklist_found = None  # Default case if no checklist category is matched
        
        if checklist_found:
            # Extract role
            role = 'student' if 'Student' in col else 'supervisor'
            # Formulate a unique key for counting
            role_key = f"{checklist_abbr}_{role}"
            if role_key not in counter_dict:
                counter_dict[role_key] = 1
            else:
                counter_dict[role_key] += 1
            # New name construction
            new_name = f"{checklist_abbr}_MC{counter_dict[role_key]}_{role}"
            renameDict[col] = new_name
        else:
            # Retain original name if no checklist category is found
            renameDict[col] = col
    # Use the rename dictonary to update column names
    df = df.rename(columns=renameDict)
    return df, renameDict

def renameColumnsRubric(df, renameDict={}):
    # Dictionary mapping the full phrases to abbreviations
    scale_map = {
        'Professionalism Scale': 'PS',
        'Communication Scale': 'CS',
        'Time Management Scale': 'TS',
        'Entrustment Scale': 'ES'
    }
    
    # Build a new rename dictionary by iterating over columns and applying transformations
    for col in df.columns:
        for key, abbreviation in scale_map.items():
            if key in col:
                renameDict[col] = abbreviation
                break  # Stop looking once the first match is found

    # Rename columns using the dictionary
    df.rename(columns=renameDict, inplace=True)
    return df, renameDict

def renameColumns(df):
    renameDict = {}
    # Rename columns for subject-specific MC columns
    df, _ = renameColumnsCode(df)
    # Rename columns for other MC columns
    df, _ = renameColumnsOther(df)
    # Rename columns for rubric questions
    df, _ = renameColumnsRubric(df)
    return df

# This is the second way of renaming columns
def notValidMCPrint(parts, col):
    if len(parts) < 2:
        print('No a valid MC column:', col, len(parts), parts)
        return True

def getMCColumnParts(col):
    
    # Checklist map for miscellaneous MC
    checklistMap = variableUtils.checklistMap

    # First split the code based on the # character
    parts = col.split('#') # first part is the code and tag while the second part is the role
    if notValidMCPrint(parts, col):
        return None
    
    if parts[0] in list(checklistMap.keys()): # for miscellaneous MC
        code = checklistMap[parts[0]]
        tag = None
    else:
        codeAndTag = re.split(r'\s+', parts[0]) # split the code and tag
        # print(codeAndTag)
        if notValidMCPrint(codeAndTag, col):
            return None
        code = codeAndTag[0]
        tag = ' '.join(codeAndTag[1:]).upper() if len(codeAndTag) > 1 else None
        if tag == 'INFIL':
            tag = 'INFILTRATION'
        if tag == 'CINIC':
            tag = 'CLINIC'

        
        # secondTag = ' '.join(codeAndTag[2:]) if len(codeAndTag) > 2 else None
        # validSecondTag = ['U/S', 'H/S', 'Remin', 'Cariostatic', 'Chart', 'Per Tooth', 'Cvek', 'Pulpot', 'Access', 'Work Length', 'Cone fit', 'Root flg']
        # if secondTag in validSecondTag:
            # tag = f"{tag}-{secondTag}"
        
            # check if tag is a valid tag
        # if tag.split('-')[0] not in variableUtils.VALID_TAGS:
            # tag = None
    
    roleAndQNo = parts[1].split('_') # split the role and question number
    if notValidMCPrint(roleAndQNo, col):
        return None
    role =  roleAndQNo[0]
    if role == '1':
            role = 'student'
    elif role == '2':
            role = 'supervisor'
    else:
        print('Invalid role:', role)
    
    if roleAndQNo[-1] == 'TEXT':
        label = 'Tooth Number'
        # return f"{code}_Tooth Number ({tag})" if tag is not None else f"{code}_Tooth Number"
    else:        
        qNo = roleAndQNo[1]
        label = f'MC{qNo}'
    return f"{code}_{label}_{role} ({tag})" if tag is not None else f"{code}_MC{qNo}_{role}"

def printDuplicateValues(renameDict):
    # Reverse the dictionary to group keys by their values
    reverseDict = {}
    for key, value in renameDict.items():
        if value in reverseDict:
            reverseDict[value].append(key)
        else:
            reverseDict[value] = [key]
    
    # Check for duplicates and print them
    duplicatesFound = False
    for value, keys in reverseDict.items():
        if len(keys) > 1:
            duplicatesFound = True
            print(f"Duplicate value: '{value}' found for keys: {keys}")
    
    if not duplicatesFound:
        print("No duplicate values found.")

def renameColumnsHeader(df, renameDict={}, renameFile = None):
    """
    Renames columns in a DataFrame based on specific conditions.

    Args:
        df (pandas.DataFrame): The DataFrame containing the columns to be renamed.

    Returns:
        pandas.DataFrame: The DataFrame with renamed columns.

    """
    if renameFile is not None:
        with open(renameFile, 'r') as f:
            renameDict = json.load(f)
            df = df.rename(columns=renameDict)
            return renameDict
    rubricMap = {
        'Professionalism': 'PS',
        'Communication': 'CS',
        'Time management': 'TS',
        'Entrustment scale': 'ES'
    }
    df.columns = [col.replace('\xa0', ' ') for col in df.columns]
    for col in df.columns:
        # if '\xa0' in col:
        #     col = col.replace('\xa0', ' ')
        renameTags = getMCColumnParts(col)
        if renameTags is not None:
            renameDict[col] = renameTags
            print(col, renameDict[col])
        elif col in rubricMap.keys():
            renameDict[col] = rubricMap[col]
        
    # In all columns replace the \xa0 with a space
    
    df = df.rename(columns=renameDict)
    for col in df.columns:
        if '631' in col:
            print(col)
    printDuplicateValues(renameDict)
    return df, renameDict

def createRenameDict(df, renameFile = variableUtils.columnRenameFile):
    df, renameDict = renameColumnsHeader(df)
    # save the dict to a json file
    with open(renameFile, 'w') as f:
        json.dump(renameDict, f)
    return df, renameDict


# Now for the overall reports functions

def getThisServiceCols(df, serviceCols):
    newServiceCols = []
    for col in serviceCols:
        if col in df.columns:
            newServiceCols.append(col)
    return newServiceCols


def mergeAndDeleteOneColumn(df, col1, col2, newCol):
    """
    Merges two columns in a DataFrame and deletes the original columns.

    Args:
        df (pandas.DataFrame): The DataFrame containing the columns.
        col1 (str): The name of the first column to merge.
        col2 (str): The name of the second column to merge.
        newCol (str): The name of the new column to create.

    Returns:
        pandas.DataFrame: The DataFrame with the merged column and the original columns dropped.
    """
    # check if the columns exist in the DataFrame
    if col1 not in df.columns or col2 not in df.columns:
        return df
    # Create the new column by merging col1 and col2
    df[newCol] = df[col1].fillna('') + ', ' + df[col2].fillna('')
    # Remove trailing comma if one column is empty or NaN
    df[newCol] = df[newCol].str.strip(', ')
    # Drop the original columns
    df = df.drop(columns=[col1, col2])
    return df

def mergeColumns(df: pd.DataFrame, serviceColMerge: list):
    for cols, new_col in serviceColMerge:
        if len(cols) == 2:
            df = mergeAndDeleteOneColumn(df, cols[0], cols[1], new_col)
        else:
            raise ValueError("Each tuple must contain exactly two columns to merge.")
    return df

def findMCColumns(df, code=None, role=None, tag = None):
    # Regex pattern to match 'MC' followed by a digit, specific code, and role
    if code is None and role is None:
        pattern = rf'MC\d+'
    elif code is None:
        pattern = rf'MC\d+_{role}'
    elif role is None:
        pattern = rf'{code}_MC\d+'
    else:
        pattern = rf'{code}_MC\d+_{role}'
    # print(f"Pattern: {pattern}")
    if tag is not None:
        pattern = pattern + rf'\s?\({tag}\)?'
    # Filter columns using the regex pattern
    print(f'MC search pattern: {pattern}')
    matched_columns = [col for col in df.columns if re.search(pattern, str(col))]    
    return matched_columns


def extractCodes(service_str: str, location: str):
    """
    Extracts restorative codes from a service string and associates them with a location.

    Args:
        service_str (str): The service string containing restorative codes.
        location (str): The location where the service is performed ('Simulation' or 'CLINIC').

    Returns:
        list: A list of tuples containing the extracted restorative codes and their associated location.

    Example:
        >>> extractRestorativeCodes('D123 D456', 'Simulation')
        [('123', 'SIM'), ('456', 'SIM')]
    """
    
    if pd.isna(service_str):
        return []
    

    
    # Extract all the codes from the service string
    codeRawList = service_str.split(',')
    # print(codeRawList)
    # print(codeRawList)
    codes = []

    
    for string in codeRawList:
        # if location is not None and 'Simulation' in str(location):
        #     tag = 'SIM'
        # else:
        #     tag = 'CLINIC'
        if string is not None:
            try:
                tag = string.split(' ')[1]
                if tag not in ['SIM', 'CLINIC']:
                    tag = 'SIM' if 'Simulation' in str(location) else 'CLINIC'
            except:
                # print('No tag found in:', string)
                tag = 'SIM' if 'Simulation' in str(location) else 'CLINIC'
        code = re.findall(r'\b(\d+)\b', string)
        for c in code:

            if len(c) <= 2:
                continue
            elif code == '941':
                continue
            elif c=='011':
                tag = 'COE'
            elif c=='012':
                tag = 'POE'
            elif c=='013':
                if 'Paeds Specific' in string:
                    tag = 'PAEDS SPECIFIC'
                tag = 'LIMITED OE'
            # WORRY ABOUT  Paeds Specific as well

            elif c== '014':
                tag = 'CONSULTATION'
            elif c == '022':
                tag = 'I/O RAD'
            elif c== '061':
                tag = 'VITALITY'
            elif c == '071':
                tag = 'DIAGNOSTIC MODEL'
            elif c == '072':
                tag = "PHOTOS"
            elif c == '074':
                tag = 'PHYSICAL MODELS'
            # elif c == '114' or c == '115':
            #     if 'Hand' in string:
            #         tag = 'H/S'
            #     elif ''
            elif c == '121':
                if 'remin' in string:
                    tag = tag + ' REMIN'
                elif 'cariostatic' in string:
                    tag = tag + ' CARIOSTATIC'
            elif c == '416':
                if '+2' in string:
                    tag = tag + ' +2'
                elif '+3' in string:
                    tag = tag + ' +3'
                else:
                    tag = tag + ' RCT ADD'
            elif c == '711' or c == '712' or c== '713' or c=='714':
                if 'Finish' in string:
                    tag = 'FINISH'
                elif 'Primary' in string:
                    tag = 'PRIMARY'
                elif 'Secondary' in string:
                    tag = 'SECONDARY'
                elif 'Occlusal' in string:
                    tag = 'OCCLUSAL'
                elif 'Try-in' in string:
                    tag = 'TRY-IN'
                elif 'review' in string:
                    tag = 'REVIEW'
            # worry about secondary and occlusal later
            elif c == '737':
                tag = 'RESILIENT LINING'
            elif c == '741':
                tag = 'ADJUSTMENT'
            elif c == '743':
                tag = 'RELINING'
            elif c== '744':
                if 'insert' in string:
                    tag = 'RELINING INSERT'
                elif 'impression' in string:
                    tag = 'RELINING PARTIAL'
            elif c == '414':
                if 'Cvek' in string:
                    tag = tag + ' CVEK'
                elif 'Deciduous' in string:
                    tag = tag + ' DECIDUOUS' 
                else:
                    print(f'No Cvek or Deciduous found in {string}')
    
            elif c == '415' and tag == 'SIM':
                if 'Access' in  string:
                    tag = tag + ' ACCESS'
                elif 'Working length' in string:
                    tag = tag + ' WORK LENGTH'
                else:
                    print(f'No Access or Working length found in {string}')
    
            elif (c == '417' or c == '418') and tag == 'SIM':
                if 'Cone fit' in string:
                    tag = tag + ' CONE FIT'
                elif 'Root filling' in string:
                    tag = tag + ' ROOT FLG'
                else:
                    print(f'No Cone fit or Root flag found in {string}')
            elif c == '579' and tag == 'SIM':
                tag = 'BONDING'

            elif (c=='587' or c=='586') and tag == 'SIM':
                if '[Separators]' in string:
                    tag = tag + ' SEPARATORS'
                else:
                    tag = tag + ' TOOTH PREP'
                
            # print(string, c, tag)
            codes.append((c, tag))
    return codes

# Function to find checklist columns
def findChecklistColumns(df, checklistMap):
    # Construct the regex pattern for checklist columns
    pattern = re.compile(rf"({'|'.join(checklistMap.values())})_MC\d+")
    # Print the constructed pattern for debugging
    # print(f"Constructed Pattern: {pattern}")
    # Filter columns using the regex pattern
    # print(df.columns)
    matched_columns = [col for col in df.columns if pattern.match(col)]    
    return matched_columns


def convertRubricScale(df, rubricQues):
    for col in rubricQues:
        df[col] = df[col].str.extract(r'Lvl (\d+)')[0]
        df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).astype('Int64')
    return df

 
def extractGeneralServiceCode(serviceStr):
    if pd.isna(serviceStr):
        return []
    
    # Regular expression to extract the code and location
    pattern = r'(\b[A-Z]+\b).*?\((.*?)\)'
    matches = re.findall(pattern, serviceStr)
    return matches

def getCodeDict(df, serviceCols, colClinicChoice, colServiceGeneral):
    print("\n Getting the code dictionary-----------------------------------------------------------S")
    codeDict = {}
    # Iterate over each row in the DataFrame
    for idx, row in df.iterrows():
        # Get codes from 'Restorative Services'
        # print(idx)
        allCodes = []

        # Get codes from the service columns
        for col in serviceCols:
            if col not in df.columns:
                pprint(f"Column {col} not found in the DataFrame")
                continue
            
            allCodes += extractCodes(row[col], row[colClinicChoice]) # Find all the codes done for a row
            print(idx, row[col], row[colClinicChoice], allCodes)
        if colServiceGeneral in df.columns:                           # Find general codes and modifiers like Infiltration or Block
            # print(f"General service column found: {row[colServiceGeneral]}")
            general_codes = extractGeneralServiceCode(row[colServiceGeneral])
            if len(general_codes) > 0:
                pprint(f"General codes: {general_codes}")
            # make the second element of the tuple upper case
            general_codes = [(code, location.upper()) for code, location in general_codes]
            allCodes += general_codes

        for codeTuple in allCodes: # A tuple with (code, location)
            code = codeTuple[0]
            modifier = codeTuple[1] if len(codeTuple) > 1 else None
            key = f"{code}_{modifier}" if modifier else code  # Create a key like '533_CLINIC' or 'LA'

            if key not in codeDict:
                codeDict[key] = []

            codeDict[key].append(idx)
    
    # pprint(codeDict) # A dictionary with keys as codes and values as indices of rows where the code was done
    return codeDict


# vectorize the rubricQues
def vectoriseColumn(columnName, df, maxColValue, newRubricQues: set):
    df[columnName] = df[columnName].fillna(0).astype(int)
    for i in range(1, maxColValue + 1):
        df[f'{columnName}-{i}'] = (df[columnName] >= i).astype(int)
        newRubricQues.add(f'{columnName}-{i}')

def vectoriseRubricQues(df, rubricQues, newRubricQues):
    # vectorise the rubricQues
    for col in rubricQues:
        # print(df[col])
        maxColValue = int(df[col].max())
        # print(col, maxColValue)
        vectoriseColumn(col, df, maxColValue, newRubricQues)
        # df.drop(columns=[col], inplace=True)
    return df

def splitDfByCodes(df, serviceCols, colClinicChoice, colServiceGeneral, beforeCols: list, mcColumns: list, rubricQues: list, afterCols: list):
    serviceCols = getThisServiceCols(df, serviceCols) # get the columns that are present in the DataFrame
    afterCols = getThisServiceCols(df, afterCols) # get the columns that are present in the DataFrame
    codeDict = getCodeDict(df, serviceCols, colClinicChoice, colServiceGeneral) # get the dictionary of codes and indices
    
    # Filtering columns for each specific code
    for key in codeDict:

        code, modifier = key.split('_') if '_' in key else (key, None) # modifier is the location like SIM or CLINIC    
        # if code != '587':
        #     continue
        relevantColumns = []
        
        # print(df.columns)
        for col in df.columns:
            if '\xa0' in col:
                print('Found \\xa0 in column name:', col)
                col = col.replace('\xa0', ' ')

            # pprint(col)
            if code in col or f'{code}_Tooth Number' in col:

                # print(f'Code found in {col}')
                # if '(SIM)' in col or '(CLINIC)' in col or '(Infiltration)' in col or '(Block)' in col: # check for modifier only if SIM or CLINIC is present
                if '(' in col:  # If there is a modifier in the column name
                    # print(f'Modifier found in {col}')
                    # if code == '587':
                    #     if modifier.upper() == col:
                    #         relevantColumns.append(col)
                    
                   if modifier in col:
                        relevantColumns.append(col)
                else:
                    relevantColumns.append(col)
        
        checklistCols = findChecklistColumns(df, variableUtils.checklistMap)
        if len(checklistCols) > 0:
            relevantColumns += checklistCols

        print(f'Code: {code}, Modifier: {modifier}')
        print(relevantColumns)
        # Combine relevant subject columns and general columns
        selectedColumns = beforeCols + relevantColumns + rubricQues + afterCols
        # Create a DataFrame for each code using selected columns
        codeDict[key] = df.loc[codeDict[key], selectedColumns].copy()
    
    # codeDict['All'] = df[beforeCols + mcColumns + rubricQues +  afterCols]

    # Convert rubric scales to integers
    # newRubricQues = set()
    for key, dftest in codeDict.items():
        dftest = convertRubricScale(dftest, rubricQues)
        # df = vectoriseRubricQues(dftest, rubricQues, newRubricQues)
    return codeDict, rubricQues


# Now to save the data to Excel
def getColumnLetter(col_idx):
        column = ""
        while col_idx > 0:
            col_idx, remainder = divmod(col_idx - 1, 26)
            column = chr(65 + remainder) + column
        return column

def saveDf(df, path, sheet_name, numColsToColor = None): 
    # Check if the file exists
    try:
        with pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Define a format object for red color fill with XlsxWriter.
            red_format = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            font_color = openpyxl.styles.Font(color='CF2D06')
            ce524_format = openpyxl.styles.PatternFill(start_color='D7BDE2', end_color='D7BDE2', fill_type='solid')  # Purple

            # Apply the format based on a conditional rule (cell value == 1).
            for idx, col in enumerate(df.columns, 1):
                if numColsToColor is not None and idx > numColsToColor:
                    break
                column_letter = getColumnLetter(idx)
                for cell in worksheet[f'{column_letter}2:{column_letter}{len(df) + 1}']:
                    for c in cell:
                        if c.value == 1:
                            c.fill = red_format
                            c.font = font_color
    except FileNotFoundError:
        with pd.ExcelWriter(path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]

            # Define a format object for red color fill with XlsxWriter.
            red_format = openpyxl.styles.PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
            font_color = openpyxl.styles.Font(color='CF2D06')
            ce524_format = openpyxl.styles.PatternFill(start_color='D7BDE2', end_color='D7BDE2', fill_type='solid')  # Purple

            # Apply the format based on a conditional rule (cell value == 1).
            for idx, col in enumerate(df.columns, 1):
                if numColsToColor is not None and idx > numColsToColor:
                    break
                column_letter = getColumnLetter(idx)
                for cell in worksheet[f'{column_letter}2:{column_letter}{len(df) + 1}']:
                    for c in cell:
                        if c.value == 1:
                            c.fill = red_format
                            c.font = font_color


def getWeighted(df, colsmc, colstm, colses, colsps, colscs, colsCE, dfCalc=None):
    if dfCalc is None:
        dfCalc = df.copy()
    # print(dfCalc.shape, df.shape, len(colsmc), len(colstm), len(colses), len(colsps), len(colscs))
    weights = [0.8, 0.05, 0.05, 0.05, 0.05]

    # Calculate the sum for each column group
    mcSum = dfCalc[colsmc].sum(axis=1, skipna=True)
    tmSum = dfCalc[colstm].sum(axis=1, skipna=True)
    esSum = dfCalc[colses].sum(axis=1, skipna=True)
    psSum = dfCalc[colsps].sum(axis=1, skipna=True)
    csSum = dfCalc[colscs].sum(axis=1, skipna=True)

    # Count non-NA values for each column group
    mcCount = dfCalc[colsmc].notna().sum(axis=1).replace(0, 1)
    tmCount = dfCalc[colstm].notna().sum(axis=1).replace(0, 1)
    esCount = dfCalc[colses].notna().sum(axis=1).replace(0, 1)
    psCount = dfCalc[colsps].notna().sum(axis=1).replace(0, 1)
    csCount = dfCalc[colscs].notna().sum(axis=1).replace(0, 1)

    # Calculate the weighted score
    df['Weighted'] = (
        (mcSum * weights[0] / mcCount.replace(0, np.nan)) + 
        (tmSum * weights[1] / tmCount.replace(0, np.nan)) + 
        (esSum * weights[2] / esCount.replace(0, np.nan)) + 
        (psSum * weights[3] / psCount.replace(0, np.nan)) + 
        (csSum * weights[4] / csCount.replace(0, np.nan))
    ) * 100
    
    df[f'Weighted'] = df[f'Weighted'].apply(lambda x: round(x, 2))
    # Penalize for critical errors by reducing the weighted score by 10%
    df['Weighted_penalty_10']= df[f'Weighted'] * df[colsCE].apply(lambda x: 0.9 if x == 'Yes' else 1.0)
    df['Weighted_penalty_10'] = df['Weighted_penalty_10'].apply(lambda x: round(x, 2))
    # Penalize for critical errors by 20%
    df['Weighted_penalty_20']= df[f'Weighted'] * df[colsCE].apply(lambda x: 0.8 if x == 'Yes' else 1.0)
    df['Weighted_penalty_20'] = df['Weighted_penalty_20'].apply(lambda x: round(x, 2))


def getColorColumns(df, code):
    colstm = df.filter(regex=f'^TS-').columns.tolist()
    colsps = df.filter(regex=f'^PS-').columns.tolist()
    colses = df.filter(regex=f'^ES-').columns.tolist()
    colscs = df.filter(regex=f'^CS-').columns.tolist()
    colsmc = findMCColumns(df, code=code)
    return colsmc + colstm + colses + colsps + colscs

def convertToGuttman(d, colId, colDate, rubricQues, workbookPath, key, colCE, mcColumns =None, colComments: list = None,
                     rubricValues = None):
    """
    Converts the given DataFrame `d` into a Guttman format.
    Parameters:
    - d (DataFrame): The input DataFrame.
    - colId (str): The column name for the student ID.
    - colDate (str): The column name for the date.
    - rubricQues (list): A list of rubric question column names.
    - workbookPath (str): The path to the workbook.
    - key (str): The key used for processing.
    - colCE (str): The column name for the CE.
    Returns:
    - None
    Raises:
    - None
    """
    pass

    # keep only MC columns, rubricQues and Student ID
    code = key.split('_')[0]
    if mcColumns is None:
        mc_columns = findMCColumns(d, code=code)
        mc_columns = [col for col in mc_columns if 'supervisor' in col.lower()]
        if code == '115' or code == '114':
            newMCcolumns = mc_columns
        else:
            pattern = rf'({code}_MC\d+)'
            new_column_names = {}
            counter = {}
            for col in mc_columns:
                match = re.search(pattern, col)
                if match:
                    newCol = match.group(1)
                    # Ensure the new column name is unique
                    if newCol in new_column_names.values():
                        if newCol not in counter:
                            counter[newCol] = 1
                        counter[newCol] += 1
                        newCol = f"{newCol}+{counter[newCol]}"
                    new_column_names[col] = newCol
            d.rename(columns=new_column_names, inplace=True)
            newMCcolumns = list(new_column_names.values())
    else:
        mc_columns = mcColumns
        newMCcolumns = mc_columns
            
    pprint(mc_columns)
    newrubricQues = set()
    d = vectoriseRubricQues(d, rubricQues, newrubricQues)
    newrubricQues = list(newrubricQues)
    
    # this only if we import rubricQues 
    # for col in rubricQues:
    #     if col in d.columns:
    #         newrubricQues.append(col)
    checklistCols = findChecklistColumns(d, variableUtils.checklistMap)
    checklistCols = [col for col in checklistCols if 'supervisor' in col.lower()]
    toothcol =[col for col in d.columns if 'Tooth Number' in col]
    toothcolsupervisor = [col for col in toothcol if 'supervisor' in col.lower()]
    print(f'Tooth found {toothcol} {toothcolsupervisor}')
    columns = [colId, colDate, variableUtils.colNameG, variableUtils.colNameF] + newMCcolumns + newrubricQues + checklistCols + variableUtils.rubricQues + toothcolsupervisor+ variableUtils.afterCols
    if colComments is not None:
        columns += colComments
    newCols = []

    for col in columns:
        if col in d.columns:
            newCols.append(col)
        else:
            print(f"Column {col} not found in the DataFrame")
    columns = newCols
    # All the tooth number columns

    print(columns)
    d = d[columns]
    d = d.loc[:, ~d.columns.duplicated()]
    colstm = d.filter(regex=f'^TS-').columns.tolist()
    colsps = d.filter(regex=f'^PS-').columns.tolist()
    colses = d.filter(regex=f'^ES-').columns.tolist()
    colscs = d.filter(regex=f'^CS-').columns.tolist()

    # Replace MC values 'Not Reviewed' with 'NA' and convert 'No' and 'Yes' to 0 and 1
    d.replace('Not Reviewed', 'NA', inplace=True)
    d.replace('Not observed or not recorded', 'NA', inplace=True)
    d.replace('Not observed or recorded', 'NA', inplace=True)
    d.replace('None', np.nan, inplace=True)
    d.replace('', np.nan, inplace=True)
    if colCE in d.columns:
        temp = d[colCE].copy()
    # d[newMCcolumns] = d[newMCcolumns].replace({'No': 0, 'Yes': 1, 'Completed': 1, 'Not completed': 0})
    d.replace('No', 0, inplace=True)
    d.replace('Yes', 1, inplace=True)
    d.replace('Completed', 1, inplace=True)
    d.replace('Not completed', 0, inplace=True)
    if colCE in d.columns:
        d[colCE] = temp
    # pprint(temp)
    # Convert 'NA' to np.nan for calculation
    d_for_calc = d.replace('NA', np.nan)
    # Converting newMCcolumns to integers, non-convertible strings become NaN
    # pprint(d_for_calc)
    for col in newMCcolumns:
        # print(d_for_calc[col])
        d_for_calc[col] = pd.to_numeric(d_for_calc[col], errors='coerce')
    # Create a total score column
    calcColumns = newMCcolumns + newrubricQues
    # pprint(calcColumns)
    # pprint(d_for_calc)
    d['MC Total'] = d_for_calc[newMCcolumns].sum(axis=1, skipna=True)
    d['Grand Total'] = d_for_calc[calcColumns].sum(axis=1, skipna=True)
    
    d['Yes'] = d[newMCcolumns].apply(lambda x: (x == 1).sum() + (x=='1').sum(), axis=1)
    d['No'] = d[newMCcolumns].apply(lambda x: (x == 0).sum() + (x=='0').sum(), axis=1)
    d['Not Reviewed'] = d[newMCcolumns].apply(lambda x: (x == 'NA').sum(), axis=1)
    d['Not Filled'] = d[newMCcolumns].apply(lambda x: (x == '').sum() + x.isna().sum() + (x=='None').sum(), axis=1)
    d['Total MC items'] = len(newMCcolumns)
    
    d['MC Score'] = (d['Yes'] + variableUtils.notReviewedW*d['Not Reviewed']) / (d['Yes'] + d['No'] + variableUtils.notReviewedW*d['Not Reviewed'])
    d['MC Score'] = d['MC Score'].apply(lambda x: round(x, 2))
    
    if rubricValues is not None:
        rubricW = rubricValues['rubricW']
        rubricDenom = rubricValues['rubricDenom']
        rubricQues = rubricValues['rubricQues']
        d['Rubric Score'] = d.apply(lambda row: sum(row[ques] * rubricW[ques] / rubricDenom[ques] for ques in rubricQues), axis=1)/sum(rubricW.values())
        d['Rubric Score'] = d['Rubric Score'].apply(lambda x: round(x, 2))

        d['Total Score'] = d['MC Score'] * variableUtils.mcScoreW + d['Rubric Score'] * variableUtils.rubricScoreW
        d['Total Score'] = d['Total Score'].apply(lambda x: round(x, 2))
        
        if colCE in d.columns:
            # Add 20% penalty for critical errors
            d['Total Score CE Penalty (20%)'] = d['Total Score'] * d[colCE].apply(lambda x: 0.8 if x == 'Yes' else 1.0)
            d['Total Score CE Penalty (20%)'] = d['Total Score CE Penalty (20%)'].apply(lambda x: round(x, 2))

            # Add 10% penalty for critical errors
            d['Total Score CE Penalty (10%)'] = d['Total Score'] * d[colCE].apply(lambda x: 0.9 if x == 'Yes' else 1.0)
            d['Total Score CE Penalty (10%)'] = d['Total Score CE Penalty (10%)'].apply(lambda x: round(x, 2))

    # if colCE in d.columns:
    #     getWeighted(d, newMCcolumns, colstm, colses, colsps, colscs, colCE, d_for_calc)
    # Sort by total score
    d.sort_values(by='Grand Total', ascending=False, inplace=True)

    # Calculate the total for each column and add as a new row
    column_totals = d_for_calc[calcColumns].sum(skipna=True)
    column_totals['Grand Total'] = np.nan  # Set Grand Total for 'Column Total' to NaN
    d.loc['Column Total'] = pd.Series(column_totals, index=calcColumns)
    d = d.sort_values(by='Column Total', axis=1, ascending=False)

    # save the dataframe to excel
    savepath = os.path.splitext(workbookPath)[0] + ' guttman.xlsx'
    colsColor = len(newMCcolumns) + len(colstm) + len(colses) + len(colsps) + len(colscs)
    saveDf(d, savepath, key.replace('/', '-'), colsColor)
    return d



def checkAttendence(workbookPath, cohort=None, studentListPath='2024 MDS Student List_v10.xlsx'):
    """
    Check the attendance of students in a given workbook.
    Parameters:
    - workbookPath (str): The path to the workbook file.
    - studentListPath (str): The path to the student list file. Default is '2024 MDS Student List_v10.xlsx'.
    - cohort (str): The cohort to filter the student list. Default is None.
    Returns:
    None
    Prints:
    - Students Attended: A list of unique student IDs who attended.
    - Students in Cohort: A list of unique student IDs in the specified cohort.
    - Students who did not attend: A list of student IDs who did not attend.
    """

    colCohort = variableUtils.colCohort
    colId = variableUtils.colId
    df = pd.read_excel(workbookPath)
    # df = loadDfFromSheet(workbook, 'Sheet0')
    # df = removeFirstRow(df)
    
    # Get list of students
    studentsAttended = df[colId].unique().astype(pd.Int64Dtype)
    print('Students Attended: ')
    pprint(studentsAttended)
    
    # Get list of students from student list
    studentDf = pd.read_excel(studentListPath)
    # selectionTupleList = [(colCohort, 'DDS2 (2024)')]
    studentDf = getDfbyColumnValue(studentDf, colCohort, cohort)
    students = studentDf[colId].unique().astype(pd.Int64Dtype)
    print('\nStudents in Cohort: ')
    pprint(students)
    # Get list of students who did not attend
    print('\nStudents who did not attend: ')
    for student in students:
        if student not in studentsAttended:
            print(student)


def getMCValueCounts(sdf, code, tag, folderPath, cohort=None):
    """
    Get value counts for each Marking Checklist column.
    """
    if cohort is not None:
        df = sdf[sdf['Cohort'] == cohort]
    else:
        df = sdf.copy()    
    # Define the colors and order for the values
    colors = {'Yes': 'green', 'No': 'red', 'Not Reviewed': 'lightgrey', 'Not Filled': 'white'}
    order = ['Yes', 'No', 'Not Reviewed', 'Not Filled']
    # Remove rows with NaN or empty Student ID
    df = df.dropna(subset=['Student ID'])
    df = df[df['Student ID'] != '']
    df['Student ID'] = df['Student ID'].astype('Int64').astype(str)

    # Identify MC columns
    mc_columns = [col for col in df.columns if 'MC' in col and 'Total' not in col]
    
    # Create a DataFrame to store the counts for each MC column
    counts_df = pd.DataFrame(index=mc_columns, columns=['Yes', 'No', 'Not Reviewed', 'Not Filled'])

    # Count the values for each MC column
    for col in mc_columns:
        counts_df.at[col, 'Yes'] = (df[col] == 1).sum()
        counts_df.at[col, 'No'] = (df[col] == 0).sum()
        counts_df.at[col, 'Not Reviewed'] = (df[col] == 'NA').sum()
        counts_df.at[col, 'Not Filled'] = (df[col] == '').sum() + df[col].isna().sum()

    # Plot the stacked bar chart
    fig, ax = plt.subplots(figsize=(12, 8))
    
    bottom = None
    for value in order:
        ax.bar(counts_df.index, counts_df[value], label=value, color=colors[value], edgecolor='black', bottom=bottom)
        if bottom is None:
            bottom = counts_df[value]
        else:
            bottom += counts_df[value]

    ax.set_xlabel('Marking Checklist columns')
    ax.set_ylabel('Count')
    title = f'Marking Checklist values {code} ({tag}) {cohort}' if cohort is not None else f'Marking Checklist values {code} ({tag})'
    ax.set_title(title)
    ax.legend(title='Values', loc='upper right', bbox_to_anchor=(1.2, 1))
    plt.xticks(rotation=90)
    savepath = f'{folderPath}/{code}_{tag}_{cohort}_MC_columns.png' if cohort is not None else f'{folderPath}/{code}_{tag}_MC_columns.png'
    plt.savefig(savepath, bbox_inches='tight')
    plt.show()


def getImportance(df, code, tag, folderPath, cohort=None):
    """
    Get the feature importances for the Random Forest Regressor model.
    """
    colCohort = variableUtils.colCohort
    rubricQues = variableUtils.rubricQues
    # If there are duplicate columns keep one of them
    df = df.loc[:, ~df.columns.duplicated()]
    if cohort is not None:
        newDf = df[df[colCohort]==cohort]
    else:
        newDf = df.copy()
    newDf.replace({'Yes': 1, 'No': 0, 'Not Assessed': np.nan, 'Not Reviewed': np.nan, 'Completed': 1, 'Not completed': 0, 'NA': np.nan}, inplace=True)
    display(newDf.head())
    display(newDf.columns)
    mc_columns_test = findMCColumns(newDf)
    # mc_columns_test = [col for col in mc_columns_test if 'supervisor' in col]
    newDf = newDf[mc_columns_test + rubricQues+ ['MC Total']]
    newDf[mc_columns_test] = newDf[mc_columns_test].replace('', pd.NA)
    newDf[mc_columns_test] = newDf[mc_columns_test].astype(pd.Int64Dtype())
    print(mc_columns_test)
    colmcTotal = 'MC Total'
    newDf[colmcTotal] = newDf[colmcTotal].replace('', pd.NA)
    newDf[colmcTotal] = newDf[colmcTotal].astype(pd.Int64Dtype())
    colmcTotalPossible = 'MC total possible'
    # newDf[colmcTotal] = newDf[mc_columns_test].sum(axis=1, skipna=True).astype(pd.Int64Dtype())
    newDf[colmcTotalPossible] = newDf[mc_columns_test].count(axis=1)
    newDf = newDf[(newDf[colmcTotalPossible]>5)]
    # display(newDf)
    colmcPercent= 'MC %'
    newDf[colmcPercent] = (newDf[colmcTotal]/newDf[colmcTotalPossible]*100)
    # display(newDf)
    newDf.to_csv(f'{folderPath}\\{code}.csv')
    
    # Drop the rows with missing values in the target column
    newDf = newDf.dropna(subset=[colmcPercent])
    # Split the data into training and testing sets

    X = newDf[mc_columns_test]
    y = newDf[colmcPercent]
    if len(y) < 5:
        print(f'Not enough data for {code} ({tag})')
        return

    # for col in rubricQues:
    #     newDf2 = newDf[newDf[col].notnull()]
    #     X = newDf2[mc_columns_test]
    #     y= newDf2[col]
    imputer = SimpleImputer(strategy='mean')
    X_imputed = imputer.fit_transform(X)
    X_train, X_test, y_train, y_test = train_test_split(X_imputed, y, test_size=0.2, random_state=42)

    # Train a Random Forest Regressor
    model = RandomForestRegressor(n_estimators=100, random_state=42)
    model.fit(X_train, y_train)

    # Predict and evaluate the model
    y_pred = model.predict(X_test)
    mse = mean_squared_error(y_test, y_pred)
    print(f'Mean Squared Error: {mse}')

    # Get feature importances
    feature_importances = model.feature_importances_
    feature_names = X.columns

    # Create a DataFrame for visualization
    importance_df = pd.DataFrame({
        'Feature': feature_names,
        'Importance': feature_importances
    }).sort_values(by='Importance', ascending=False)

    # Plot the feature importances
    plt.figure(figsize=(10, 8))
    plt.barh(importance_df['Feature'], importance_df['Importance'])
    plt.xlabel('Importance')
    plt.ylabel('Feature')
    title = f'Feature Importances for {code} ({tag}) {cohort}' if cohort is not None else f'Feature Importances for {code} ({tag})'
    plt.title(title)
    plt.gca().invert_yaxis()
    plt.tight_layout()
    savepath = f'{folderPath}/{code}_{tag}_{cohort}_FeatureImportances.png' if cohort is not None else f'{folderPath}/{code}_{tag}_FeatureImportances.png'
    plt.savefig(savepath, bbox_inches='tight')
    plt.show()

    # Get counts of each type of value
    getMCValueCounts(df, code, tag, folderPath, cohort=cohort)



def getValueCountsStudent(sdf, code, role, folderPath, cohort=None):
    """
    Get student wise value counts for the MC columns in the DataFrame.
    """
    order = ['Yes', 'No', 'Not Reviewed', 'Not Filled']
    colors = {'Yes': 'blue', 'No': 'orange', 'Not Reviewed': 'lightgrey', 'Not Filled': 'white'}
    sdf = sdf.dropna(subset=['Student ID'])
    # drop empty string
    sdf = sdf[sdf['Student ID'] != '']
    sdf['Student ID'] = sdf['Student ID'].astype('Int64').astype(str)
    if cohort is not None:
        df = sdf[sdf['Cohort'] == cohort]
    else:
        df = sdf.copy()

    # Identify MC columns
    mc_columns = [col for col in df.columns if 'MC' in col and 'Total' not in col]
    
    # Create a DataFrame to store the counts for each student
    counts_df = pd.DataFrame(index=df.index, columns=['Yes', 'No', 'Not Reviewed', 'Not Filled'])
    # display(df)
    # Count the values for each student in the MC columns
    counts_df['Yes'] = df[mc_columns].apply(lambda x: (x == 1).sum(), axis=1)
    counts_df['No'] = df[mc_columns].apply(lambda x: (x == 0).sum(), axis=1)
    counts_df['Not Reviewed'] = df[mc_columns].apply(lambda x: (x == 'NA').sum(), axis=1)
    counts_df['Not Filled'] = df[mc_columns].apply(lambda x: (x == '').sum() + x.isna().sum(), axis=1)
    counts_df['Student ID'] = df['Student ID']

    # pprint(counts_df)
    # Plot the stacked bar chart
    width = len(counts_df)*1.1
    fig, ax = plt.subplots(figsize=(14, 8))

    bottom = None
    for value in order:
        ax.bar(counts_df['Student ID'], counts_df[value], label=value, color=colors[value], edgecolor='black', bottom=bottom)
        if bottom is None:
            bottom = counts_df[value]
        else:
            bottom += counts_df[value]

    ax.set_xlabel('Student ID')
    ax.set_ylabel('Count')
    ax.set_title(f'Marking Checklist values counts for {code} ({role})')
    ax.legend(title='Values', loc='upper right', bbox_to_anchor=(1.2, 1))
    plt.xticks(rotation=90)
    plt.savefig(f'{folderPath}/{code}_{role}.png', bbox_inches='tight')
    plt.show()


def autopct(pct, total):
    """
    Generate the autopct string for a pie chart.
    Parameters:
    - pct (float): The percentage value of the data point.
    - total (int): The total value of the data points.
    Returns:
    - str: The formatted autopct string.
    Example:
    >>> autopct(25, 100)
    '25%\n(25)'
    Usage: lambda pct: autopct(pct, total)
    """
    
    val = int(round(pct * total / 100.0))
    return '{:.0f}% ({v:d})'.format(pct, v=val) if pct > 0 else ''



def anonymize_column(column):
    unique_values = column.dropna().unique()  # Get unique non-null values
    mapping = {str(value): idx for idx, value in enumerate(unique_values, start=1)}  # Create a mapping
    reverse_mapping = {idx: value for value, idx in mapping.items()}  # Reverse mapping for reference
    return column.map(mapping), mapping, reverse_mapping

def custom_agg(series):
    """
    Custom aggregation function for a Series. Marking Checklist values
    """
    unique_vals = series.dropna().unique()
    if len(unique_vals) == 0: # All values are NA
        return pd.NA
    if len(unique_vals) == 1: # Only one unique value 
        return unique_vals[0]
    else:
        if '1' in unique_vals:
            return '1'
        elif 1 in unique_vals:
            return 1
        elif '0' in unique_vals:
            return '0'
        elif 0 in unique_vals:
            return 0
        else:
            print(f"No unique value found in {unique_vals}")
            return pd.NA


def getPairCounts(df, colPairBy = variableUtils.colId, colPairWith = variableUtils.colSupervisor):
    # Create a dictionary to count pairs
    pair_counts = {}

    # Group by Student ID
    grouped = df.groupby(colPairBy)

    # Loop through each group
    for _, group in grouped:
        # Get unique CE Names for each student
        ce_names = group[colPairWith].unique()
        
        # Get all combinations of pairs (should be only one pair per student in this setup)
        pairs = list(combinations(sorted(ce_names), 2))
        
        for pair in pairs:
            if pair in pair_counts:
                pair_counts[pair] += 1
            else:
                pair_counts[pair] = 1

    # Convert the dictionary to a DataFrame
    pairs_df = pd.DataFrame(pair_counts.items(), columns=['Pair', '# of Pairs'])

    # Display the result
    display(pairs_df)
    return pairs_df

def aggregator(dfGuttman, mcCols, colCE=None, colCEReason=None, colComments: list = None, colSupervisor = None):
    print("\nAggregating data...")
    rubricQues = [ques for ques in variableUtils.rubricQues if ques in dfGuttman.columns]
    aggFuncs = {col: 'first' for col in dfGuttman.columns if col not in mcCols + rubricQues}
    aggFuncs.update({col: custom_agg for col in mcCols})
    aggFuncs.update({col: 'max' for col in rubricQues})
    matching_columns = [col for col in dfGuttman.columns if any(pattern in col for pattern in variableUtils.newRubricQuesPatterns)]
    aggFuncs.update({col: 'max' for col in matching_columns})
    if colCE is not None:
        aggFuncs.update({colCE: lambda x: 'Yes' if 'Yes' in x.values else 'No'})
    if colCEReason is not None:
        dfGuttman[colCEReason] = dfGuttman[colCEReason].fillna(' ').astype(str)
        aggFuncs.update({colCEReason: lambda x: ', '.join(x.values).strip(' ,')})
    if colComments is not None:
        for col in colComments:
            dfGuttman[col] = dfGuttman[col].fillna(' ').astype(str)
            aggFuncs.update({col: lambda x: '\n\n '.join(x.values).strip(' \n')})
    if colSupervisor is not None:
        dfGuttman[colSupervisor] = dfGuttman[colSupervisor].fillna(' ').astype(str)
        aggFuncs.update({colSupervisor: lambda x: ', '.join(x.values).strip(' ,')})
    dfTemp = dfGuttman.groupby([variableUtils.colId, variableUtils.colDate], as_index=False).agg(aggFuncs)  # aggregate the data
    dfTemp = dfTemp.reindex(columns=dfGuttman.columns, fill_value=None)  # reindex the columns
    return dfTemp



def createTable(df, title, colRatio:list, tableWidth = 0.9, customTextCols = [], 
            tableTextStyle = variableUtils.tableTextStyle, topPadding = 12, bottomPadding = 12, cellHighlight = False, headerColor = '#9C27B0', titleStyle = variableUtils.subsubheadingStyle):
    print(f'Creating table for {title}')
    if df.empty:
        table = Paragraph("No data found", variableUtils.subsubheadingStyle)
    else:
        data = [df.columns.to_list()] + df.values.tolist()
        
        # Convert the custom text columns to paragraphs
        for i in range(1, len(data)):
            for j in customTextCols:
                data[i][j] = Paragraph(str(data[i][j]), tableTextStyle)
        
        if colRatio is not None:
            colWidths = [ratio/sum(colRatio) * variableUtils.pageSize[0] * tableWidth for ratio in colRatio]
        else:
            colWidths = [1 for i in range(len(df.columns))] # Equal column widths
        # print(f'Column widths: {colWidths}')
        table = Table(data, colWidths=colWidths)
        # print(data)
        table_style = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(headerColor)),  # Header row
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FFFFFF')),  # Header text
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),  # Center align all cells
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),  # Center align all cells
            ('GRID', (0, 0), (-1, -1), 1, colors.black),  # Add border around cells
            # ('ALIGN', (3, 1), (3, -1), 'LEFT'),  # Left align Reason column cells
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),  # Change font to bold
            ('FONTSIZE', (0, 0), (-1, -1), 14),  # Increase font size
            ('BOTTOMPADDING', (0, 0), (-1, -1), bottomPadding),  # Increase bottom padding
            ('TOPPADDING', (0, 0), (-1, -1), topPadding),  # Increase top padding
        ])
        table.setStyle(table_style)

    mergedElement = KeepTogether([Paragraph(title, titleStyle), Spacer(1, 6), table, Spacer(1, 12)])

    # Add red colour where cell values are No
    if not cellHighlight:
        return mergedElement
    if df.empty:
        return mergedElement
    for i in range(1, len(data)):
        for j in range(len(data[i])):
            if data[i][j] == 'No':
                table.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.red)]))
            if data[i][j] == 'Yes':
                table.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.green)]))
    return mergedElement

def createSplitTable(df, title, colRatio:list, tableWidth=0.9, customTextCols=[], 
                     tableTextStyle=variableUtils.tableTextStyle, topPadding=12, bottomPadding=12, 
                     cellHighlight=False, headerColor='#9C27B0', titleStyle=variableUtils.subsubheadingStyle):
    
    print(f'Creating split table for {title}')
    
    if df.empty:
        table = Paragraph("No data found", variableUtils.subsubheadingStyle)
        mergedElement = KeepTogether([Paragraph(title, titleStyle), Spacer(1, 6), table, Spacer(1, 12)])
        return mergedElement

    else:
        data = [df.columns.to_list()] + df.values.tolist()

        # Convert custom text columns to Paragraphs
        for i in range(1, len(data)):
            for j in customTextCols:
                data[i][j] = Paragraph(str(data[i][j]), tableTextStyle)

        if colRatio is not None:
            colWidths = [ratio/sum(colRatio) * variableUtils.pageSize[0] * tableWidth/2 for ratio in colRatio]
        else:
            colWidths = [1 for _ in range(len(df.columns))]

        # Split rows
        headerRow = data[0]
        bodyRows = data[1:]
        splitPoint = (len(bodyRows) + 1) // 2  # +1 for safe split if odd number

        leftData = [headerRow] + bodyRows[:splitPoint]
        rightData = [headerRow] + bodyRows[splitPoint:]

        # Create left and right tables
        leftTable = Table(leftData, colWidths=colWidths)
        rightTable = Table(rightData, colWidths=colWidths)

        tableStyle = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(headerColor)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#FFFFFF')),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 14),
            ('BOTTOMPADDING', (0, 0), (-1, -1), bottomPadding),
            ('TOPPADDING', (0, 0), (-1, -1), topPadding),
        ])

        leftTable.setStyle(tableStyle)
        rightTable.setStyle(tableStyle)

        # Now combine left and right tables into one row with two columns
        combinedTable = Table(
            [[leftTable, rightTable]],
            colWidths=[variableUtils.pageSize[0]*tableWidth/2]*2,
            hAlign='CENTER',
                style=[
        ('VALIGN', (0, 0), (-1, -1), 'TOP')  # This line is critical
    ]
        )

    mergedElement = KeepTogether([Paragraph(title, titleStyle), Spacer(1, 6), combinedTable, Spacer(1, 12)])

    # Add cellHighlight if needed
    if not cellHighlight:
        return mergedElement

    # Coloring Yes/No
    for i in range(1, len(leftData)):
        for j in range(len(leftData[i])):
            if isinstance(leftData[i][j], str) and leftData[i][j] == 'No':
                leftTable.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.red)]))
            if isinstance(leftData[i][j], str) and leftData[i][j] == 'Yes':
                leftTable.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.green)]))
    for i in range(1, len(rightData)):
        for j in range(len(rightData[i])):
            if isinstance(rightData[i][j], str) and rightData[i][j] == 'No':
                rightTable.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.red)]))
            if isinstance(rightData[i][j], str) and rightData[i][j] == 'Yes':
                rightTable.setStyle(TableStyle([('TEXTCOLOR', (j, i), (j, i), colors.green)]))

    return mergedElement


def createPlotImage(fig):
        buf = BytesIO()
        fig.savefig(buf, format='png', bbox_inches='tight')
        buf.seek(0)
        return buf

def addPlotImage(fig, ratio = None):
        plotImage = createPlotImage(fig)
        pageSize = variableUtils.pageSize
        topMargin = variableUtils.topMargin
        bottomMargin = variableUtils.bottomMargin
        leftMargin = variableUtils.leftMargin
        rightMargin = variableUtils.rightMargin
        image = Image(plotImage)
        # print(image.drawWidth, image.drawHeight)
        
        # Resize image to fit within margins
        max_height = pageSize[1] - topMargin - bottomMargin  # Max height for Page
        max_width = pageSize[0] - leftMargin - rightMargin  # Max width for Page
        if ratio is not None:
            max_width = max_width * ratio
            max_height = max_height * ratio
        aspect_ratio = min(max_width / image.drawWidth, max_height / image.drawHeight)
        image.drawWidth *= aspect_ratio
        image.drawHeight *= aspect_ratio
        # print(image.drawWidth, image.drawHeight, aspect_ratio)
        # if idx + 1 < numSubplots:
        #    self.elements.append(PageBreak())
        #self.elements.append(PageBreak())
        plt.close(fig)
        return(image)

def cleanEntry(codeList):
    validCodes = []
    for entry in codeList:
        if isinstance(entry, str) and 'LA' in entry.upper():
            validCodes.append('LA')
        
        matches = re.findall(r'(\d+)(?:\s*[xX]\s*(\d+))?', str(entry))

        first = True
        for code, multiplier in matches:
            if len(code) == 1:
                continue

            if len(code) > 3:
                if len(code) % 3 != 0:
                    code = '0' + code
                splitCodes = [code[i:i+3] for i in range(0, len(code), 3)]
            else:
                if first and len(code) == 2:
                    code = '0' + code
                splitCodes = [code] if len(code) == 3 else []

            count = int(multiplier) if multiplier else 1
            for splitCode in splitCodes:
                if len(splitCode) == 3:
                    validCodes.extend([splitCode] * count)
            
            first = False
    return validCodes

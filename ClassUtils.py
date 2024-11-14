from Utils import *
from variableUtils import *
import variableUtils
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import warnings
import openpyxl
import os
from pprint import pprint
import re


class DataCleaning:
    
    def __init__(self, workbookPath):
        self.workbookPath = workbookPath

    def filterOneSheet(self, sheetName = 'Sheet0'):
        """
        Filters the first sheet of a workbook by removing fake names, removing the first row (excluding column names),
        converting a specific column to 'Int64' data type, renaming columns, and saving the filtered data to a new Excel file.

        Parameters:
        - self: The instance of the class.

        Returns:
        - None
        """
        print("Filtering data....")
        workbookPath = self.workbookPath
        folderPath, fileName, fileExt = getFolderandFileName(workbookPath)
        df = pd.read_csv(workbookPath, encoding = 'ISO-8859-1') if fileExt == '.csv' else pd.read_excel(workbookPath)
        # workbook, folderPath, fileName = loadWorkbook(workbookPath)
        # print(folderPath, fileName)
        # df = loadDfFromSheet(workbook, sheetName)
        df = removeFirstRow(df, True) # remove the first row and False to keep the first row as column names
        display(df.head())
        for i, row in df.iterrows():
            row[colId] = str(row[colId]).replace('.0', '')
            try:
                row[colId] = int(row[colId])
            except:
                print(f"Error converting {row[colId]} to int")
                continue
            # id = row[colId].astype('Int64')
        df[colId] = df[colId].astype('Int64')
        studentList = getStudentList()
        df = removeFakeNames(df, colNameG, colNameF, studentList)
        # df = renameColumns(df)
        df, _ = renameColumnsHeader(df)
        savepath = os.path.splitext(workbookPath)[0] + ' filtered.xlsx'
        folderPath = os.path.dirname(savepath)
        if not os.path.exists(folderPath):
            os.makedirs(folderPath)
        df.columns = [col.replace('\xa0', ' ') for col in df.columns]
        df.to_excel(savepath, index=False)
        print(f"Filtered data saved to {savepath}")
        return savepath


class CreateGuttman:
    def __init__(self, savePath=None, workbookPath=None):
        if savePath is None:
            cleaner = DataCleaning(workbookPath)
            savePath = cleaner.filterOneSheet()

        self.savePath = savePath
        self.workbookPath = workbookPath
        self.codesDf = None

        # self.workbook, self.folder, self.fileName = loadWorkbook(self.savePath)
        # self.df = loadDfFromSheet(self.workbook, sheetName='Sheet1') # load the first sheet
        self.folder, self.fileName, self.fileExt = getFolderandFileName(self.savePath)
        self.df = pd.read_excel(self.savePath, sheet_name='Sheet1')
        print("Data loaded successfully!!!!")
        
        self.df = mergeColumns(self.df, serviceColMerge) # merge columns service
        self.df = mergeColumns(self.df, [([colSupervisorChoice, colSupervisorOther], colSupervisor)])
        print("Columns merged successfully!!!!")
        
        print('\n Converting column Id to Int64......')
        self.df[colId] = self.df[colId].astype('Int64')
        
        print('\n Converting column Date to datetime......')
        print(self.df[colDate].unique())
        self.df[colDate] = self.df[colDate].apply(convertDate)
        # print("\nDates converted to datetime format")
        display(self.df[colDate].unique())

        print("\nRemoving students with no valid ID.....")
        studentList = getStudentList()
        for i, row in self.df.iterrows():
            if row[colId] not in studentList:
                print(f"Removing student with ID: {row[colId]}")
                self.df.drop(i, inplace=True)

    
    def getCodesDf(self, df, selectionTupleList: list = []):
        rubricQues = variableUtils.rubricQues
        colId = variableUtils.colId
        colCE = variableUtils.colCE
        df[colId] = df[colId].astype('Int64')
        # df[colDate] = pd.to_datetime(df[colDate], format = '%d/%m/%Y').dt.date
        # Remove None column names
        mc_columns = findMCColumns(df)
        # pprint(mc_columns)
        x= df.copy()

        if len(selectionTupleList) != 0:
            for col, value in selectionTupleList:
                x = getDfbyColumnValue(x, col, value) # get the dataframe with the selected values
        # print(rubricQues)
        codesDf, rubricQues = splitDfByCodes(x, serviceCols, colClinicChoice, colServiceGeneral, beforeCols, mc_columns, rubricQues, afterCols)
        return codesDf

    
    def saveSplitDf(self, selectionTupleList: list = []):
        # convert date
        # remove students with no valid ID
        print('Splitting the data')
        self.codesDf = self.getCodesDf(self.df, selectionTupleList)
        pprint(f'Codes Df dict: {self.codesDf.keys()}')
        # sort the keys
        self.codesDf = dict(sorted(self.codesDf.items()))
        for key, x in self.codesDf.items():
            # print(x.columns)
            if '941' in key:
                continue
            # if key != 'LA_Infiltration':
            #     continue
            print('\n saving DF to file for item: ', key)
            x.columns = x.columns.str.strip()
            # remove columns with all NaN values
            # x = x.dropna(axis=1, how='all')
            # for col in x.columns:
            #     print(col)
            path = self.workbookPath if self.workbookPath is not None else self.savePath
            savepath = os.path.splitext(path)[0] + ' split.xlsx'
            saveDf(x, savepath, key.replace('/', '-'))
        return savepath
    
    
    def createGuttman(self, savepath=None, alreadyExistingGuttmanPath = None):
        if savepath is None:
            savepath = self.saveSplitDf()

        workbook, folder, fileName = loadWorkbook(savepath)
        existingWorkbook = None
        existingSheetNames = []

        print(self.workbookPath)
        folder, fileName2, _ = getFolderandFileName(self.workbookPath)

        if alreadyExistingGuttmanPath is not None:
            existingWorkbook, _, _ = loadWorkbook(alreadyExistingGuttmanPath)
            existingSheetNames = existingWorkbook.sheetnames
        print("Already existing sheet names: ", existingSheetNames)
        rubricQues = variableUtils.rubricQues
        colId = variableUtils.colId
        colCE = variableUtils.colCE
        colDate = variableUtils.colDate

        labelBlank = 'Not Filled'
        otherCols = [colComplex, colClinicType, colDate, colSupervisor, colFinished, colCE, colCEReason]
        labelRubric = 'Rubric Score'
        dfTemplate = pd.DataFrame(columns=[colId, 'Yes', 'No', 'Not Reviewed', labelBlank, 
                                           'Total MC items'] + rubricQues + ['MC Score', labelRubric] + otherCols +['Item', 
                                                                                'Total Score','Total Score CE Penalty (20%)', 'Total Score CE Penalty (10%)'])
        countsDfAll = dfTemplate.copy()
        for sheet in workbook.sheetnames:
            print(sheet)
            code = sheet.split('_')[0]
            if sheet in existingSheetNames:
                print(f"{sheet} already exists in the workbook")
                continue
            
            df = loadDfFromSheet(workbook, sheetName=sheet)
            mcCols = findMCColumns(df, code)
            path = self.workbookPath if self.workbookPath is not None else self.savePath
            
            # Create rubricW, rubricDenom and rubricQues
            rubricQues = variableUtils.rubricQues
            rubricQues = [i for i in rubricQues if i in df.columns]            # only get the rubricQues that are in the df
            # rubricDenom = {i: df[i].max() for i in rubricQues}                # Get max value of each rubricQues as the denominator
            
            # check if the df is clinic or sim
            countsClinic = df[colClinicChoice].value_counts()
            highestClinic = countsClinic.idxmax()
            if 'Simulation' in highestClinic:
                print("Simulation data")
                rubricW = {'PS': 0.05, 'CS': 0.05, 'TS': .1, 'ES': .1}
                rubricDenom = {'PS': 2, 'CS': 2, 'TS': 4, 'ES': 4}
            else:
                print("Clinic data")
                rubricW = {'PS': 0.1, 'CS': 0.1, 'TS': .1, 'ES': .1}
                rubricDenom = {'PS': 4, 'CS': 4, 'TS': 4, 'ES': 4}
            
            # remove the items in rubricW that are not in rubricQues
            rubricW = {k: v for k, v in rubricW.items() if k in rubricQues}

            # Finish the dictionary
            rubricValues = {'rubricW': rubricW, 'rubricDenom': rubricDenom, 'rubricQues': rubricQues}

            dfGuttman = convertToGuttman(df, colId, colDate, rubricQues, path, sheet, colCE, None, None, rubricValues)
            mcCols = findMCColumns(dfGuttman, code)
            dfTemp = aggregator(dfGuttman, mcCols, colCE, colCEReason, None)
            print(f"Saving {sheet} to {folder}/{fileName2} best.xlsx")
            dfBest = convertToGuttman(dfTemp, colId, colDate, rubricQues, f'{folder}/{fileName2} best.xlsx', sheet, colCE, mcCols, None, rubricValues)

            countsDf = dfTemplate.copy()
            for col in countsDf.columns:
                if col in dfBest.columns:
                    # except the last row
                    countsDf[col] = dfBest[col][:-1]
            countsDf['Item'] = sheet    
            countsDfAll = pd.concat([countsDfAll, countsDf], ignore_index=True)
        
        countsDfAll.sort_values(by=[colId], inplace=True)
        countsDfAll.to_excel(f'{folder}/{fileName2} marks.xlsx', index=False)

        print(countsDfAll[colId].value_counts())

        # Aggregate the data on Student ID, take average of Total Score
        df = countsDfAll.copy()
        df[colId] = df[colId].astype('Int64')
        aggFuncs = {col: 'first' for col in df.columns if col not in [colId, 'Item']}
        aggFuncs.update({'Total Score': 'mean', 'Total Score CE Penalty (20%)': 'mean', 'Total Score CE Penalty (10%)': 'mean'})
        df = df.groupby([colId], as_index=False).agg(aggFuncs)
        df['Total Score'] = df['Total Score'].apply(lambda x: round(x, 2))
        df['Total Score CE Penalty (20%)'] = df['Total Score CE Penalty (20%)'].apply(lambda x: round(x, 2))
        df['Total Score CE Penalty (10%)'] = df['Total Score CE Penalty (10%)'].apply(lambda x: round(x, 2))
        df = df[[colId, 'Total Score', colCE, colCEReason, 'Total Score CE Penalty (20%)', 'Total Score CE Penalty (10%)']]
        df.sort_values('Total Score', ascending=False, inplace=True)
        df.to_excel(f'{folder}/{fileName2} marks aggregated.xlsx', index=False)
            

        

